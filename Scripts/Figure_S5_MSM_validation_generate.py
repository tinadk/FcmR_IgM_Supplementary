#!/usr/bin/env python3
"""
Figure S5 – MSM validation (implied timescales, CK test, eigenvalue spectrum,
          timescale separation). Single combined 2×2 panel.
Reads data and parameters from config.py.  No titles.  Subplot labels A–D
placed at the top-left outside of each panel.
Exports: ITS, CK, eigenvalues, and separate timescales CSV, plus combined Excel.
"""

import sys
from pathlib import Path
ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from config import OUT_DIR, DPI, MPL_RCPARAMS

import numpy as np
import matplotlib.pyplot as plt
import pandas as pd

plt.rcParams.update(MPL_RCPARAMS)

# ---------- Paths ----------
MSM_DIR = ROOT / 'MSM' / 'archive' / 'complete' / 'data' / 'msm'
OUTPUT_DIR = OUT_DIR / 'Figure_S5_MSM_validation'
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ---------- Load data ----------
its = np.load(MSM_DIR / 'its_k150.npy')
lags = np.load(MSM_DIR / 'lags_k150.npy') * 0.1          # frames → ns
its_ns = its * 0.1

ck = np.load(MSM_DIR / 'ck_errors.npy')
lags_ck = (np.arange(1, ck.shape[0] + 1) * 0.1).astype(float)

eig = np.real(np.load(MSM_DIR / 'eigenvalues.npy'))
spec = eig[eig < 0.9999]                                 # discard trivial eigenvalue ~1
t_scale = -1.0 / np.log(spec[:20]) * 0.1                 # first 20 timescales (ns)

# ---------- CSV export ----------
# Implied timescales
its_header = 'lag_ns,' + ','.join(f'ITS{i+1}' for i in range(its_ns.shape[1]))
its_csv = np.column_stack([lags, its_ns])
np.savetxt(OUTPUT_DIR / 'Figure_S5_ITS.csv', its_csv, delimiter=',',
           header=its_header, comments='', fmt='%.6f')

# CK errors
ck_header = 'lag_ns,' + ','.join(f'CK_error_set{i+1}' for i in range(ck.shape[1]))
ck_csv = np.column_stack([lags_ck, ck])
np.savetxt(OUTPUT_DIR / 'Figure_S5_CK.csv', ck_csv, delimiter=',',
           header=ck_header, comments='', fmt='%.6f')

# Eigenvalues and timescales (full, NaN padded)
timescale_padded = np.pad(t_scale, (0, len(spec) - len(t_scale)),
                          constant_values=np.nan)
eig_csv = np.column_stack([np.arange(1, len(spec) + 1), spec, timescale_padded])
np.savetxt(OUTPUT_DIR / 'Figure_S5_eigenvalues.csv', eig_csv, delimiter=',',
           header='index,eigenvalue,timescale_ns', comments='', fmt='%.6f')

# Separate timescales CSV (first 20 only)
timescale_only_csv = np.column_stack([np.arange(1, len(t_scale) + 1), t_scale])
np.savetxt(OUTPUT_DIR / 'Figure_S5_timescales.csv', timescale_only_csv,
           delimiter=',', header='index,timescale_ns', comments='', fmt='%.6f')

# ---------- Excel export (four sheets) with formatting ----------
excel_path = OUTPUT_DIR / 'Figure_S5_MSM_validation_data.xlsx'
try:
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        # Prepare dataframes with capitalized column headers
        its_df = pd.DataFrame(its_ns, columns=[f'ITS{i+1}' for i in range(its_ns.shape[1])])
        its_df.insert(0, 'lag_ns', lags)
        its_df.columns = [col[0].upper() + col[1:] for col in its_df.columns]
        its_df.to_excel(writer, sheet_name='ITS', index=False)

        ck_df = pd.DataFrame(ck, columns=[f'CK_error_set{i+1}' for i in range(ck.shape[1])])
        ck_df.insert(0, 'lag_ns', lags_ck)
        ck_df.columns = [col[0].upper() + col[1:] for col in ck_df.columns]
        ck_df.to_excel(writer, sheet_name='CK', index=False)

        eig_df = pd.DataFrame({
            'index': np.arange(1, len(spec) + 1),
            'eigenvalue': spec,
            'timescale_ns': timescale_padded
        })
        eig_df.columns = [col[0].upper() + col[1:] for col in eig_df.columns]
        eig_df.to_excel(writer, sheet_name='Eigenvalues', index=False)

        ts_df = pd.DataFrame({
            'index': np.arange(1, len(t_scale) + 1),
            'timescale_ns': t_scale
        })
        ts_df.columns = [col[0].upper() + col[1:] for col in ts_df.columns]
        ts_df.to_excel(writer, sheet_name='Timescales', index=False)

    # Apply formatting: font Arial, alignment left + top
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment
    wb = load_workbook(excel_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(name='Arial')
                cell.alignment = Alignment(horizontal='left', vertical='top')
    wb.save(excel_path)
    print(f'Excel saved to {excel_path} with Arial font and left-top alignment')
except Exception as e:
    print(f'Excel export failed (but CSV files are still available): {e}')

# ---------- Combined 2×2 figure ----------
fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(14, 10),
                                             constrained_layout=True)

# Top‑left: Implied timescales
for i in range(min(5, its_ns.shape[1])):
    ax1.plot(lags, its_ns[:, i], marker='o', label=f'ITS {i+1}')
ax1.set_xlabel('Lag time (ns)')
ax1.set_ylabel('Timescale (ns)')
ax1.legend(fontsize=9)
ax1.text(-0.12, 1.06, 'A', transform=ax1.transAxes,
         fontweight='bold', fontsize=9, va='bottom', ha='center', color='black')

# Top‑right: Chapman–Kolmogorov test
for i in range(ck.shape[1]):
    ax2.plot(lags_ck, ck[:, i], marker='o', label=f'Set {i+1}')
ax2.set_xlabel('Lag time (ns)')
ax2.set_ylabel('CK error')
ax2.legend(fontsize=9)
ax2.text(-0.12, 1.06, 'B', transform=ax2.transAxes,
         fontweight='bold', fontsize=9, va='bottom', ha='center', color='black')

# Bottom‑left: Eigenvalue spectrum
ax3.plot(range(1, len(spec) + 1), spec, 'o-', markersize=4)
ax3.set_xlabel('Eigenvalue index')
ax3.set_ylabel('Eigenvalue (real part)')
ax3.text(-0.12, 1.06, 'C', transform=ax3.transAxes,
         fontweight='bold', fontsize=9, va='bottom', ha='center', color='black')

# Bottom‑right: Timescale separation
ax4.plot(range(1, len(t_scale) + 1), t_scale, 's-', color='crimson')
ax4.set_xlabel('Eigenvalue index')
ax4.set_ylabel('Implied timescale (ns)')
ax4.text(-0.12, 1.06, 'D', transform=ax4.transAxes,
         fontweight='bold', fontsize=9, va='bottom', ha='center', color='black')

# Save figure
for ext in ('png', 'jpg', 'pdf'):
    plt.savefig(OUTPUT_DIR / f'Figure_S5_MSM_validation.{ext}',
                dpi=DPI, bbox_inches='tight', facecolor='white')
plt.close()

print(f'Figure_S5 and all data files saved to {OUTPUT_DIR}')